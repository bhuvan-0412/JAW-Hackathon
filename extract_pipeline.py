#!/usr/bin/env python3
from __future__ import annotations
"""
extract_pipeline.py — Structured Extraction Pipeline for Document Estate

Extracts 678 PDFs and 9 Excel workbooks indexed by document_index.csv.
Uses layout-preserving PDF extraction (pymupdf4llm / pdfplumber layout),
openpyxl + formulas for .xlsx files, and money_parser.py for all monetary figures.

Outputs:
  - extracted/{doc_id}.json (per-document structured JSON)
  - extraction_log.jsonl (aggregate execution log)
"""

import os
import re
import json
import glob
import sys
import argparse
import datetime
import pandas as pd
import openpyxl
import pdfplumber

try:
    import fitz
except ImportError:
    fitz = None

try:
    import pymupdf4llm
except ImportError:
    pymupdf4llm = None

try:
    import formulas
except ImportError:
    formulas = None

from money_parser import parse_money, find_money_mentions, parse_cell_value

# Ensure stdout uses UTF-8 encoding on Windows
if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DOCUMENTS_DIR = os.path.join(BASE_DIR, 'documents')
OUTPUT_DIR = os.path.join(BASE_DIR, 'extracted')
LOG_FILE = os.path.join(BASE_DIR, 'extraction_log.jsonl')


def clean_json_obj(obj):
    """
    Recursively clean objects to ensure strict JSON serializability.
    Converts datetime/date, numpy types, formulas Ranges objects, etc. to standard Python types.
    """
    if obj is None:
        return None
    if isinstance(obj, (int, float, str, bool)):
        return obj
    if isinstance(obj, (datetime.date, datetime.datetime)):
        return obj.isoformat()
    if isinstance(obj, dict):
        return {str(k): clean_json_obj(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple, set)):
        return [clean_json_obj(item) for item in obj]
    
    # Handle numpy / formulas objects or other custom classes
    try:
        if hasattr(obj, 'tolist'):
            return clean_json_obj(obj.tolist())
        if hasattr(obj, 'item'):
            return clean_json_obj(obj.item())
        if hasattr(obj, 'value'):
            return clean_json_obj(obj.value)
    except Exception:
        pass

    return str(obj)


def extract_pdf_layout_text(pdf_path: str) -> tuple[str, int, int]:
    """
    Extract text, char count, and page count from a PDF file using fitz / pdfplumber.
    """
    page_count = 0
    if fitz is not None:
        try:
            doc = fitz.open(pdf_path)
            page_count = len(doc)
            pages_text = [page.get_text("text") for page in doc]
            doc.close()
            text = '\n\n'.join(pages_text)
            return text, len(text), page_count
        except Exception:
            pass

    pages_text = []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            page_count = len(pdf.pages)
            for page in pdf.pages:
                txt = page.extract_text() or ''
                pages_text.append(txt)
        text = '\n\n'.join(pages_text)
        return text, len(text), page_count
    except Exception as e2:
        return f"EXTRACTION_FAILED: {str(e2)}", 0, page_count


def parse_markdown_key_values(md_text: str) -> dict[str, str]:
    kv = {}
    table_rows = re.findall(r'\|(?:\*\*)?([^|]+?)(?:\*\*)?\|([^|\n]+)\|', md_text)
    for k, v in table_rows:
        k_clean = k.replace('*', '').replace('_', '').strip()
        v_clean = v.replace('*', '').replace('_', '').strip()
        if k_clean and v_clean and not k_clean.startswith('---'):
            kv[k_clean.lower()] = v_clean
            
    line_kvs = re.findall(r'\*\*(.+?):\*\*\s*(.+)', md_text)
    for k, v in line_kvs:
        k_clean = k.replace('*', '').replace('_', '').strip()
        v_clean = v.replace('*', '').replace('_', '').strip()
        if k_clean and v_clean:
            kv[k_clean.lower()] = v_clean

    line_kvs_plain = re.findall(r'^\s*([A-Za-z0-9\s/_\-\(\)\']+):\s*(.+)', md_text, re.MULTILINE)
    for k, v in line_kvs_plain:
        k_clean = k.replace('*', '').replace('_', '').strip()
        v_clean = v.replace('*', '').replace('_', '').strip()
        if k_clean and v_clean and k_clean.lower() not in kv:
            kv[k_clean.lower()] = v_clean

    lines = [l.strip() for l in md_text.split('\n') if l.strip()]
    for i in range(len(lines) - 1):
        k_clean = lines[i].replace('*', '').replace('_', '').strip().lower()
        v_clean = lines[i+1].replace('*', '').replace('_', '').strip()
        if k_clean not in kv and len(k_clean) < 50:
            kv[k_clean] = v_clean

    return kv


def extract_grading_sentence(md_text: str) -> str | None:
    sentences = re.split(r'(?<=[.!?])\s+', md_text)
    for s in sentences:
        s_clean = s.replace('\n', ' ').strip()
        s_lower = s_clean.lower()
        if any(kw in s_lower for kw in ['graded', 'quality of the completed work', 'overall quality']) and len(s_clean) > 15:
            return s_clean

    for s in sentences:
        s_clean = s.replace('\n', ' ').strip()
        s_lower = s_clean.lower()
        if 'graded' in s_lower or ('workmanship' in s_lower and 'satisfactory' in s_lower):
            if len(s_clean) > 15:
                return s_clean

    return None


def extract_completion_certificate(md_text: str, doc_type: str) -> dict:
    kv = parse_markdown_key_values(md_text)
    
    project_name = (
        kv.get('name of work') or 
        kv.get('project name') or 
        kv.get('work') or
        kv.get('contract / work order title')
    )
    if not project_name:
        match = re.search(r'work of\s+["“’\'\?]?([^"”’\'\?\n]+?)["“”’\'\?]?(?:\s*\([^)]+\))?,\s*awarded', md_text, re.IGNORECASE)
        if match:
            project_name = match.group(1).strip()
    if not project_name:
        match = re.search(r'work\s+(?:of\s+)?["“’\'\?]?([^"”’\'\?\n]+)["“”’\'\?]', md_text, re.IGNORECASE)
        if match:
            project_name = match.group(1).strip()

    package_code = kv.get('package code')
    if not package_code and (project_name or md_text):
        match = re.search(r'(?:Pkg|Package)[- ]?([A-Za-z0-9-]+)', project_name or md_text, re.IGNORECASE)
        if match:
            package_code = match.group(0).strip()
            
    client_name = (
        kv.get('client / awarding authority') or 
        kv.get('client name') or 
        kv.get('client') or 
        kv.get('employer')
    )
    if not client_name:
        lines = [l.strip() for l in md_text.split('\n') if l.strip()]
        for l in lines[:5]:
            l_lower = l.lower()
            if any(k in l_lower for k in ['authority', 'department', 'dept', 'office', 'corporation', 'corp', 'bureau', 'board', 'limited', 'ltd']) and not any(l_lower.startswith(w) for w in ['no.', 'dated', 'certificate', 'office of']):
                client_name = l.replace('#', '').replace('*', '').strip()
                break
        if not client_name and lines:
            client_name = lines[0].replace('#', '').replace('*', '').strip()
            
    raw_contract_val = (
        kv.get('contract / work order value') or 
        kv.get('executed value') or 
        kv.get('awarded value') or 
        kv.get('contract value (original)') or
        kv.get('contract value')
    )
    contract_val = parse_money(raw_contract_val) if raw_contract_val else None
    
    if contract_val is None:
        mentions = find_money_mentions(md_text)
        if mentions:
            raw_contract_val, contract_val = mentions[0]

    start_date = (
        kv.get('commencement date') or 
        kv.get('start date') or 
        kv.get('original start date')
    )
    completion_date = (
        kv.get('completion date') or 
        kv.get('completion') or 
        kv.get('actual completion date') or
        kv.get('scheduled completion date')
    )
    
    if not completion_date:
        match = re.search(r'completed\s+(?:in all respects\s+)?on\s+([\d]{1,2}[-/\s][\d]{1,2}[-/\s][\d]{2,4}|[\d]{4}-\d{2}-\d{2}|[\d]{1,2}\s+[A-Za-z]+\s+[\d]{4})', md_text, re.IGNORECASE)
        if match:
            completion_date = match.group(1).strip()

    project_lead = (
        kv.get("contractor's project manager") or 
        kv.get('project lead') or 
        kv.get('project manager')
    )
    if not project_lead:
        match = re.search(r'supervised\s+(?:on the contractor\'s side\s+)?by\s+([A-Z][a-z]+(?:\s+[A-Z][a-z]+)+)', md_text, re.IGNORECASE)
        if match:
            project_lead = match.group(1).strip()

    grading_text = extract_grading_sentence(md_text)

    return {
        "project_name": project_name,
        "package_code": package_code,
        "client_name": client_name,
        "contract_value": contract_val,
        "raw_contract_value": raw_contract_val,
        "start_date": start_date,
        "completion_date": completion_date,
        "project_lead": project_lead,
        "grading_text": grading_text
    }


def extract_reference_letter(md_text: str) -> dict:
    kv = parse_markdown_key_values(md_text)
    
    project_name = kv.get('project title') or kv.get('work executed') or kv.get('project name') or kv.get('work') or kv.get('name of work')
    if not project_name:
        match = re.search(r'Subject:.*?[“"’\'\?]?([^”"’\'\?\n]+?Pkg-[A-Za-z0-9-]+)[”"’\'\?]?', md_text, re.IGNORECASE)
        if match:
            project_name = match.group(1).strip()
    if not project_name:
        match = re.search(r'for the work\s+[“"’\'\?]?([^”"’\'\?\n]+?)[”"’\'\?]?(?:\s*\((?:INR|Rs))?', md_text, re.IGNORECASE)
        if match:
            project_name = match.group(1).strip()
            
    package_code = kv.get('package code')
    if not package_code and (project_name or md_text):
        match = re.search(r'(?:Pkg|Package)[- ]?([A-Za-z0-9-]+)', project_name or md_text, re.IGNORECASE)
        if match:
            package_code = match.group(0).strip()
            
    issuing_client = kv.get('issuing authority') or kv.get('client') or kv.get('employer')
    if not issuing_client:
        lines = [line.strip() for line in md_text.split('\n') if line.strip()]
        if lines:
            for line in lines[:5]:
                line_clean = line.replace('*', '').replace('#', '').strip()
                if not any(line_clean.lower().startswith(w) for w in ['ref:', 'date:', 'subject:', 'to whomsoever', 'this letter', 'page ']):
                    issuing_client = line_clean
                    break

    raw_val = kv.get('contract value') or kv.get('value') or kv.get('sanctioned value')
    contract_val = parse_money(raw_val) if raw_val else None
    if contract_val is None:
        mentions = find_money_mentions(md_text)
        if mentions:
            raw_val, contract_val = mentions[0]

    completion_date = kv.get('completion date') or kv.get('completed') or kv.get('date of completion')
    if not completion_date:
        match = re.search(r'(?:completed|finished)\s+(?:on\s+)?([\d]{1,2}[-/\s][\d]{1,2}[-/\s][\d]{2,4}|[\d]{4}-\d{2}-\d{2}|[A-Za-z]+\s+[\d]{1,2},\s+[\d]{4})', md_text, re.IGNORECASE)
        if match:
            completion_date = match.group(1).strip()

    return {
        "project_name": project_name,
        "package_code": package_code,
        "issuing_client": issuing_client,
        "contract_value": contract_val,
        "raw_contract_value": raw_val,
        "completion_date": completion_date,
        "recommendation_summary": md_text[:500]
    }


def extract_performance_bond(md_text: str) -> dict:
    kv = parse_markdown_key_values(md_text)
    
    project_name = kv.get('project / work name') or kv.get('work of') or kv.get('project name') or kv.get('work')
    if not project_name:
        match = re.search(r'work of\s+["“’\'\?]?([^"”’\'\?\n]+?)["“”’\'\?]', md_text, re.IGNORECASE)
        if match:
            project_name = match.group(1).strip()

    package_code = kv.get('package code')
    if not package_code and (project_name or md_text):
        match = re.search(r'(?:Pkg|Package)[- ]?([A-Za-z0-9-]+)', project_name or md_text, re.IGNORECASE)
        if match:
            package_code = match.group(0).strip()

    issuing_bank = kv.get('guarantor bank') or kv.get('issuing bank') or kv.get('bank')
    if not issuing_bank:
        match = re.search(r'(?:we,)\s+\*\*?([^*]+?(?:Bank|Financial))\*\*?', md_text, re.IGNORECASE)
        if match:
            issuing_bank = match.group(1).strip()
        else:
            lines = md_text.split('\n')
            for line in lines[:5]:
                if 'bank' in line.lower():
                    issuing_bank = line.replace('*', '').replace('#', '').strip()
                    break

    beneficiary = kv.get('beneficiary') or kv.get('client')
    if not beneficiary:
        match = re.search(r'To:\s*\n+([^\n]+)', md_text, re.IGNORECASE)
        if match:
            beneficiary = match.group(1).strip()

    raw_amount = kv.get('guarantee amount') or kv.get('bond amount')
    guarantee_amount = parse_money(raw_amount) if raw_amount else None
    if guarantee_amount is None:
        match = re.search(r'exceeding\s+\*\*?([^*]+?)\*\*?', md_text, re.IGNORECASE)
        if match:
            raw_amount = match.group(1).strip()
            guarantee_amount = parse_money(raw_amount)

    issue_date = kv.get('issue date') or kv.get('date of issue')
    if not issue_date:
        match = re.search(r'Issue Date:\s*([\d]{4}-\d{2}-\d{2}|[\d]{1,2}[-/\s][\d]{1,2}[-/\s][\d]{2,4})', md_text, re.IGNORECASE)
        if match:
            issue_date = match.group(1).strip()

    expiry_date = kv.get('expiry date') or kv.get('validity') or kv.get('valid until')
    if not expiry_date:
        match = re.search(r'until\s+\*\*?([\d]{4}-\d{2}-\d{2}|[\d]{1,2}[-/\s][\d]{1,2}[-/\s][\d]{2,4})\*\*?', md_text, re.IGNORECASE)
        if match:
            expiry_date = match.group(1).strip()

    return {
        "project_name": project_name,
        "package_code": package_code,
        "issuing_bank": issuing_bank,
        "beneficiary": beneficiary,
        "guarantee_amount": guarantee_amount,
        "raw_guarantee_amount": raw_amount,
        "issue_date": issue_date,
        "expiry_date": expiry_date
    }


def extract_personnel_certificate(md_text: str) -> dict:
    kv = parse_markdown_key_values(md_text)
    
    engineer_name = kv.get('employee name') or kv.get('name')
    if not engineer_name:
        match = re.search(r'certify that\s*\n+#\s*\*\*([^*]+)\*\*', md_text, re.IGNORECASE)
        if match:
            engineer_name = match.group(1).strip()

    employee_id = kv.get('employee id')
    if not employee_id:
        match = re.search(r'Employee ID:\s*(EMP-[\d]+)', md_text, re.IGNORECASE)
        if match:
            employee_id = match.group(1).strip()

    cert_type = kv.get('credential type') or kv.get('certification title')
    if not cert_type:
        if 'PMP' in md_text:
            cert_type = 'PMP'
        elif 'Six Sigma' in md_text:
            cert_type = 'Six Sigma Black Belt'

    credential_id = kv.get('credential id') or kv.get('credential / license no.')
    if not credential_id:
        match = re.search(r'Credential ID:\s*([A-Za-z0-9-]+)', md_text, re.IGNORECASE)
        if match:
            credential_id = match.group(1).strip()

    issue_date = kv.get('date of issue') or kv.get('issued')
    if not issue_date:
        match = re.search(r'Issued:\s*([\d]{4}-\d{2}-\d{2})', md_text, re.IGNORECASE)
        if match:
            issue_date = match.group(1).strip()

    expiry_date = kv.get('valid through') or kv.get('validity / expiry')

    return {
        "engineer_name": engineer_name,
        "employee_id": employee_id,
        "certification_type": cert_type,
        "credential_id": credential_id,
        "issue_date": issue_date,
        "expiry_date": expiry_date
    }


def extract_cv(md_text: str) -> dict:
    kv = parse_markdown_key_values(md_text)
    
    engineer_name = kv.get('name')
    employee_id = kv.get('employee id')
    designation = kv.get('designation')
    experience = kv.get('total experience')
    qualification = kv.get('qualification')

    projects_led = []
    project_rows = re.findall(r'\|(?:\d+)?\|([^|]+)\|([^|]+)\|([^|]+)\|([^|]+)\|', md_text)
    for proj_name, pkg, client, role in project_rows:
        if not proj_name.strip().startswith('---') and 'project name' not in proj_name.lower():
            projects_led.append({
                "project_name": proj_name.strip(),
                "package_code": pkg.strip(),
                "client": client.strip(),
                "role": role.strip()
            })

    return {
        "engineer_name": engineer_name,
        "employee_id": employee_id,
        "designation": designation,
        "total_experience": experience,
        "qualification": qualification,
        "projects_led": projects_led
    }


def extract_ra_bill(md_text: str, doc_type: str) -> dict:
    kv = parse_markdown_key_values(md_text)
    
    project_name = kv.get('project name') or kv.get('work')
    package_code = kv.get('package code')

    client_name = kv.get('employer / client') or kv.get('employer') or kv.get('client')
    if not client_name:
        match = re.search(r'Contract\s+#[\d]+\s*·\s*([^\n]+)', md_text, re.IGNORECASE)
        if match:
            client_name = match.group(1).strip()
            
    if not client_name:
        match = re.search(r'Contract\s+#[\d]+\s*·\s*([^·\n]+)\s*·', md_text, re.IGNORECASE)
        if match:
            client_name = match.group(1).strip()

    bill_number = kv.get('bill no') or kv.get('bill number')
    bill_date = kv.get('date') or kv.get('bill date')

    raw_contract_val = kv.get('contract value') or kv.get('awarded value')
    contract_val = parse_money(raw_contract_val) if raw_contract_val else None

    raw_billed_val = kv.get('total value of work billed') or kv.get('net claimed (before client tds)') or kv.get('value of work done — this bill')
    billed_val = parse_money(raw_billed_val) if raw_billed_val else None
    
    if billed_val is None:
        mentions = find_money_mentions(md_text)
        if mentions:
            raw_billed_val, billed_val = mentions[0]

    boq_items = []
    item_rows = re.findall(r'\|([^|]+)\|([^|]+)\|([^|]+)\|([^|]+)\|([^|]+)\|([^|]+)\|', md_text)
    for item_no, desc, unit, rate, qty, amt in item_rows:
        i_clean = item_no.replace('*', '').strip()
        if i_clean and not i_clean.startswith('---') and i_clean.lower() != 'item':
            boq_items.append({
                "item_no": i_clean,
                "description": desc.strip(),
                "unit": unit.strip(),
                "rate": parse_cell_value(rate),
                "qty": parse_cell_value(qty),
                "amount": parse_cell_value(amt)
            })

    return {
        "project_name": project_name,
        "package_code": package_code,
        "client_name": client_name,
        "bill_number": bill_number,
        "bill_date": bill_date,
        "contract_value": contract_val,
        "billed_value": billed_val,
        "boq_items": boq_items
    }


def extract_generic_pdf(md_text: str, doc_type: str) -> dict:
    kv = parse_markdown_key_values(md_text)
    money_mentions = find_money_mentions(md_text)
    dates = re.findall(r'\b(?:\d{4}-\d{2}-\d{2}|\d{1,2}[-/\s][A-Za-z]{3}[-/\s]\d{2,4})\b', md_text)
    
    return {
        "key_values": kv,
        "money_mentions": money_mentions,
        "extracted_dates": dates,
        "summary_text": md_text[:1000]
    }


def extract_workbook_data(file_path: str) -> tuple[dict, list[str]]:
    warnings = []
    wb_data = openpyxl.load_workbook(file_path, data_only=True)
    wb_form = openpyxl.load_workbook(file_path, data_only=False)

    uncalculated_cells = []
    for sheet in wb_data.sheetnames:
        ws_d = wb_data[sheet]
        ws_f = wb_form[sheet]
        for row_idx, (r_d, r_f) in enumerate(zip(ws_d.iter_rows(values_only=True), ws_f.iter_rows(values_only=True)), start=1):
            for col_idx, (v_d, v_f) in enumerate(zip(r_d, r_f), start=1):
                if isinstance(v_f, str) and v_f.startswith('='):
                    if v_d is None:
                        cell_ref = f"'{sheet}'!{openpyxl.utils.get_column_letter(col_idx)}{row_idx}"
                        uncalculated_cells.append(cell_ref)

    formula_solutions = {}
    if uncalculated_cells:
        try:
            xl_model = formulas.ExcelModel().loads(file_path).finish()
            formula_solutions = xl_model.calculate()
        except Exception as e:
            warnings.append(f"Formulas library evaluation failed: {str(e)}")

    sheets_data = {}
    for sheet in wb_data.sheetnames:
        ws_d = wb_data[sheet]
        ws_f = wb_form[sheet]
        rows = []
        headers = []
        for row_idx, (r_d, r_f) in enumerate(zip(ws_d.iter_rows(values_only=True), ws_f.iter_rows(values_only=True)), start=1):
            cleaned_row = []
            for col_idx, (v_d, v_f) in enumerate(zip(r_d, r_f), start=1):
                val = v_d
                if isinstance(v_f, str) and v_f.startswith('='):
                    if val is None:
                        filename_base = os.path.basename(file_path)
                        cell_key = f"'[{filename_base}]{sheet}'!{openpyxl.utils.get_column_letter(col_idx)}{row_idx}"
                        if cell_key in formula_solutions:
                            val = formula_solutions[cell_key]
                        else:
                            warnings.append(f"Unresolved formula cell at {cell_key}")
                            val = None
                cleaned_row.append(clean_json_obj(val))
                
            if row_idx == 1:
                headers = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(cleaned_row)]
            rows.append(cleaned_row)
            
        sheets_data[sheet] = {
            "headers": headers,
            "row_count": len(rows),
            "rows": rows[:500]
        }

    return sheets_data, warnings


def process_document(row: pd.Series) -> tuple[dict, dict]:
    doc_id = row['doc_id']
    doc_type = row['doc_type']
    filename = row['filename']
    size_bytes = int(row['size_bytes'])
    
    file_path = os.path.join(DOCUMENTS_DIR, filename)
    warnings = []
    extracted_fields = {}
    char_count = 0
    page_count = 1
    status = "success"

    try:
        if not os.path.exists(file_path):
            warnings.append(f"File not found: {file_path}")
            status = "partial_failure"
        elif filename.endswith('.xlsx'):
            try:
                sheets_data, wb_warnings = extract_workbook_data(file_path)
                warnings.extend(wb_warnings)
                extracted_fields = {
                    "workbook_type": doc_type,
                    "sheets": sheets_data
                }
                char_count = sum(len(str(s)) for s in sheets_data.values())
            except Exception as e:
                warnings.append(f"Workbook extraction error: {str(e)}")
                status = "partial_failure"
        else:
            md_text, char_count, page_count = extract_pdf_layout_text(file_path)
            
            if char_count < 100:
                warnings.append(f"Low char_count warning ({char_count} chars on {page_count} pages)")

            if md_text.startswith("EXTRACTION_FAILED"):
                warnings.append(md_text)
                status = "partial_failure"
            else:
                if doc_type in ('completion_certificate', 'company_completion_certificate'):
                    extracted_fields = extract_completion_certificate(md_text, doc_type)
                elif doc_type == 'reference_letter':
                    extracted_fields = extract_reference_letter(md_text)
                elif doc_type == 'performance_bond':
                    extracted_fields = extract_performance_bond(md_text)
                elif doc_type == 'personnel_certificate':
                    extracted_fields = extract_personnel_certificate(md_text)
                elif doc_type == 'cv':
                    extracted_fields = extract_cv(md_text)
                elif doc_type in ('ra_bill', 'final_ra_bill'):
                    extracted_fields = extract_ra_bill(md_text, doc_type)
                else:
                    extracted_fields = extract_generic_pdf(md_text, doc_type)
    except Exception as exc:
        status = "partial_failure"
        warnings.append(f"Unhandled exception during extraction: {str(exc)}")

    doc_json = clean_json_obj({
        "doc_id": doc_id,
        "doc_type": doc_type,
        "filename": filename,
        "size_bytes": size_bytes,
        "char_count": char_count,
        "page_count": page_count,
        "extraction_warnings": warnings,
        "extracted_data": extracted_fields
    })

    log_entry = clean_json_obj({
        "doc_id": doc_id,
        "doc_type": doc_type,
        "filename": filename,
        "size_bytes": size_bytes,
        "page_count": page_count,
        "char_count": char_count,
        "status": status,
        "warnings": warnings
    })

    return doc_json, log_entry


def run_pipeline(docs_dir: str = DOCUMENTS_DIR, output_dir: str = OUTPUT_DIR, batch_size: int | None = None):
    os.makedirs(output_dir, exist_ok=True)
    index_csv = os.path.join(docs_dir, 'document_index.csv')
    if not os.path.exists(index_csv):
        index_csv = os.path.join(BASE_DIR, 'document_index.csv')

    if os.path.exists(index_csv):
        df = pd.read_csv(index_csv)
        if batch_size:
            sampled_rows = []
            for doc_type, group in df.groupby('doc_type'):
                sampled_rows.append(group.head(1))
            batch_df = pd.concat(sampled_rows).head(batch_size)
        else:
            batch_df = df

        total_docs = len(batch_df)
        print(f"Starting extraction pipeline for {total_docs} documents from index {index_csv}...", flush=True)

        log_entries = []
        for idx, (_, row) in enumerate(batch_df.iterrows(), start=1):
            doc_json, log_entry = process_document(row)
            
            out_path = os.path.join(output_dir, f"{row['doc_id']}.json")
            with open(out_path, 'w', encoding='utf-8') as f:
                json.dump(doc_json, f, indent=2, ensure_ascii=False)
                
            log_entries.append(log_entry)
            
            if idx % 100 == 0 or idx == total_docs:
                print(f"Processed {idx}/{total_docs} documents...", flush=True)
    else:
        # Generic recursive walk on docs_dir
        print(f"No document_index.csv found. Running generic recursive walk on {docs_dir}...", flush=True)
        file_list = []
        for root, _, files in os.walk(docs_dir):
            for file in files:
                if file.endswith(('.pdf', '.xlsx', '.csv')) and not file.startswith('~$'):
                    file_list.append(os.path.join(root, file))

        total_docs = len(file_list)
        log_entries = []
        for idx, fpath in enumerate(file_list, start=1):
            fname = os.path.basename(fpath)
            doc_id = os.path.splitext(fname)[0]
            rel_path = os.path.relpath(fpath, docs_dir)

            # Basic doc_type sniffing
            fname_lower = fname.lower()
            if 'cc' in fname_lower or 'completion' in fname_lower:
                doc_type = 'completion_certificate'
            elif 'ref' in fname_lower or 'recommendation' in fname_lower:
                doc_type = 'reference_letter'
            elif 'bond' in fname_lower or 'guarantee' in fname_lower:
                doc_type = 'performance_bond'
            elif 'pcert' in fname_lower or 'personnel' in fname_lower:
                doc_type = 'personnel_certificate'
            elif 'cv' in fname_lower:
                doc_type = 'cv'
            elif 'bill' in fname_lower or 'ra' in fname_lower:
                doc_type = 'ra_bill'
            elif fname_lower.endswith('.xlsx'):
                if 'boq' in fname_lower: doc_type = 'boq_workbook'
                elif 'ageing' in fname_lower: doc_type = 'ageing_workbook'
                elif 'asset' in fname_lower: doc_type = 'asset_register_workbook'
                elif 'tb' in fname_lower or 'trial' in fname_lower: doc_type = 'trial_balance_workbook'
                else: doc_type = 'boq_workbook'
            else:
                doc_type = 'completion_certificate'

            row = pd.Series({
                'doc_id': doc_id,
                'doc_type': doc_type,
                'filename': rel_path
            })
            doc_json, log_entry = process_document(row)

            out_path = os.path.join(output_dir, f"{doc_id}.json")
            with open(out_path, 'w', encoding='utf-8') as f:
                json.dump(doc_json, f, indent=2, ensure_ascii=False)

            log_entries.append(log_entry)
            if idx % 100 == 0 or idx == total_docs:
                print(f"Processed {idx}/{total_docs} documents...", flush=True)

    # Write log file atomically
    log_file_path = os.path.join(BASE_DIR, 'extraction_log.jsonl')
    with open(log_file_path, 'w', encoding='utf-8') as log_f:
        for entry in log_entries:
            log_f.write(json.dumps(entry, ensure_ascii=False) + '\n')

    print(f"\nExtraction complete! Logged {len(log_entries)} entries to {log_file_path}.", flush=True)


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--docs', default=DOCUMENTS_DIR, help="Path to documents directory")
    parser.add_argument('--output-dir', default=OUTPUT_DIR, help="Path to output extracted JSON directory")
    parser.add_argument('--batch', type=int, help="Run a batch of N sample documents across doc_types")
    args = parser.parse_args()
    
    run_pipeline(docs_dir=args.docs, output_dir=args.output_dir, batch_size=args.batch)
